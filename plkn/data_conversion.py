import os
from datetime import date

from models import *
from alpha_camps import *
from bravo_camps import *
from charlie_camps import *
from delta_camps import *

from openpyxl import load_workbook

def Create_Trainee(new_name, new_ic_no, new_company_id, new_index_no, #new_blok_no,
                   new_room_no, new_bed_no, new_gender, new_race_id, new_religion_id, new_age, new_phone1, new_phone2,
                   new_address1, new_address2, new_address3, new_postcode, new_city, new_state,
                   new_is_married, new_education_id, new_occupation, new_bsn_account_no, new_is_shooter,
                   new_is_left_handed, new_is_registered, new_report_in_date,
                   new_kin_name, new_relation_id, new_kin_address, new_kin_phone, new_kin_occupation):
    try:
        new_trainee = Trainees.select().where(Trainees.ic_no==new_ic_no).get()
    except Trainees.DoesNotExist:
        if new_gender != "" and new_race_id != "":
            #filled gender, & race
            #########  RULES SET HERE ##############################################################3
            #set index no based on gender
            new_index = LastIndex.select().where(LastIndex.camp==session['camp']).get()
            if new_gender == "Male":
                new_index.male = new_index.male + 2
            else:
                new_index.female = new_index.female + 2
                
            #set company, bed & room no based on race
            group = GroupPlacement.select().where(GroupPlacement.camp==session['camp']).where(GroupPlacement.race==new_race_id).get()
            if group.alpha == 0:
                group.alpha = 1
                #put into alpha
                company_placed = Companies.select().where(Companies.name=="Alpha").get()
                #put into room--------------------------------------
                if new_gender == "Male":
                    dorm = Dormitory.select().where(Dormitory.room_no==alpha_male_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_male_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_male_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_male_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                        
                else:
                    dorm = Dormitory.select().where(Dormitory.room_no==alpha_female_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_female_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_female_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==alpha_female_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
            elif group.bravo == 0:
                group.bravo = 1
                #put into bravo
                company_placed = Companies.select().where(Companies.name=="Bravo").get()
                #put into room--------------------------------------
                if new_gender == "Male":
                    dorm = Dormitory.select().where(Dormitory.room_no==bravo_male_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_male_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_male_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_male_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                        
                else:
                    dorm = Dormitory.select().where(Dormitory.room_no==bravo_female_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_female_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_female_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==bravo_female_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
            elif group.charlie == 0:
                group.charlie = 1
                #put into charlie
                company_placed = Companies.select().where(Companies.name=="Charlie").get()
                #put into room--------------------------------------
                if new_gender == "Male":
                    dorm = Dormitory.select().where(Dormitory.room_no==charlie_male_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_male_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_male_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_male_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                        
                else:
                    dorm = Dormitory.select().where(Dormitory.room_no==charlie_female_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_female_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_female_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==charlie_female_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
            else:
                group.alpha = 0
                group.bravo = 0
                group.charlie = 0
                #put into delta
                company_placed = Companies.select().where(Companies.name=="Delta").get()
                #put into room--------------------------------------
                if new_gender == "Male":
                    dorm = Dormitory.select().where(Dormitory.room_no==delta_male_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_male_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_male_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_male_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                        
                else:
                    dorm = Dormitory.select().where(Dormitory.room_no==delta_female_dorm_1(session['camp'])).get()
                    if dorm.occupancy < 25:
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25:
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_female_dorm_2(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #BS has 3 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_female_dorm_3(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
                    elif dorm.occupancy >= 25: #SR has 4 dorms/camps
                        dorm = Dormitory.select().where(Dormitory.room_no==delta_female_dorm_4(session['camp'])).get()
                        dorm.occupancy = dorm.occupancy + 1
            #########  RULES SET HERE ##############################################################

        new_trainee = Trainees.create(name = new_name, 
                                      ic_no = new_ic_no,
                                      camp = session['camp'],
                                      company = None if new_race_id == "" else company_placed,
                                      index_no = "" if new_gender == "" or new_race_id == "" else session['camp'] + "151" + '{:0>3}'.format(str(new_index.male if new_gender == "Male" else new_index.female)),
                                      #blok_no = new_blok_no,
                                      room_no = "" if new_race_id == "" else dorm.room_no,
                                      bed_no = "" if new_race_id == "" else dorm.occupancy,
                                      gender = None if new_gender == "" else new_gender,
                                      race = None if new_race_id == "" else Races.select().where(Races.id==new_race_id).get(),
                                      religion = None if new_religion_id == "" else Religions.select().where(Religions.id==new_religion_id).get(),
                                      age = new_age,
                                      phone1 = new_phone1,
                                      phone2 = new_phone2,
                                      address1 = new_address1,
                                      address2 = new_address2,
                                      address3 = new_address3,
                                      postcode = new_postcode,
                                      city = new_city,
                                      state = new_state,
                                      is_married = True if new_is_married == "True" else False,
                                      education = Educations.select().where(Educations.id==new_education_id).get(),
                                      occupation = new_occupation,
                                      bsn_account_no = new_bsn_account_no,
                                      is_shooter = True if new_is_shooter == "True" else False,
                                      is_left_handed = True if new_is_left_handed == "True" else False,
                                      is_registered = True if new_is_registered == "True" else False,
                                      report_in_date = new_report_in_date,
                                      kin_name = new_kin_name,
                                      relation = Relations.select().where(Relations.id==new_relation_id).get(),
                                      kin_address = str(new_kin_address),
                                      kin_phone = new_kin_phone,
                                      kin_occupation = new_kin_occupation,
                                      created_by = Users.select().where(Users.email==session['email']).get())
        if not new_gender == "" and not new_race_id == "": 
            new_index.save()
        if not new_race_id == "": 
            group.save()
            dorm.save()
        return True
    else:
        return False

#--------------------------------------------------------------------------
#login user
email = "admin@admin.com"

"""
excel = load_workbook(os.path.join("upload", "similajau.xlsx"))
sheet = excel.get_sheet_by_name('SENARAI PENUH - BINTULU')
camp = "SW"

excel = load_workbook(os.path.join("upload", "sgrait.xlsx"))
sheet = excel.get_sheet_by_name('Sg.Rait, Miri')
camp = "SR"

excel = load_workbook(os.path.join("upload", "junacopark.xlsx"))
sheet = excel.get_sheet_by_name('Junaco Park, sibu')
camp = "JP"

excel = load_workbook(os.path.join("upload", "bumimas.xlsx"))
sheet = excel.get_sheet_by_name('Bumimas, Sibu')
camp = "BM"
"""

excel = load_workbook(os.path.join("upload", "bukitsaban.xlsx"))
sheet = excel.get_sheet_by_name('Bukit saban betong')
camp = "BS"


session = {'email': email, 'camp': camp}
counter = 0

#for row in sheet.iter_rows(range_string="A1:AA800", row_offset=2): #similajau
for row in sheet.iter_rows(range_string="A1:AE800", row_offset=1): #others
    if row[0].value is not None:
        #      nama          no kp         jantina        agama          kaum           
        #print row[3].value, row[2].value, row[12].value, row[26].value, row[25].value #SW
        #print row[3].value, row[2].value, row[12].value, row[28].value, row[27].value #SR
        
        nama = str(row[3].value).strip()
        icno = str(row[2].value).strip()
        jantina = "Male" if str(row[12].value).strip() == "LELAKI" else "Female"
        
        if str(row[28].value).strip() == "-" or row[28].value == "" or row[28].value == None:
            agama = ""
        else:
            religion = Religions.select().where(Religions.name.contains(str(row[28].value).strip())).get()
            agama = religion.id
        
        if str(row[27].value).strip() == "TIADA" or str(row[27].value).strip() == "LAIN-LAIN" or row[27].value == "":
            kaum = ""
        else:
            bumi = str(row[27].value).strip()
            if bumi == "BUMIPUTRA" or bumi == "BUMI LL":
                bumi = "BUMIPUTERA"
            elif bumi == "MALAYU":
                bumi = "MELAYU"
            race = Races.select().where(Races.name.contains(bumi)).get()
            kaum = race.id
        
        #calculate age
        current = long(date.today().year)
        born = long(str(row[2].value)[:2]) + 1900
        age = current - born 
        
        
        #then enter data
        if Create_Trainee(nama, icno, None, None, None, None, jantina, kaum, agama, age, 
                          None, None, None, None, None, None, None, None, None, "1", None, None, None, None, None, None, None,
                          "1", None, None, None) :
            counter += 1
            print str(counter) + " Trainee Inserted"
        else:
            print "error!"
            break